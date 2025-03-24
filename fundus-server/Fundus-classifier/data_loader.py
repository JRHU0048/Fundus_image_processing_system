# 加载数据集
import torch
from torchvision import datasets, transforms
from torch.utils.data import DataLoader, random_split
import os
from torch.utils.data import DataLoader, random_split, Dataset
import xlrd
import cv2
import numpy as np
import openpyxl
from PIL import Image

# 加载OIA-ODIR数据集使用
class ODIRDataset(Dataset):
    def __init__(self, image_root, gt_filepath, transform=None):
        """
        初始化 ODIR 数据集
        参数:
            image_root (str): 图像文件所在的根目录
            gt_filepath (str): 包含真实标签的 Excel 文件路径
            transform (callable, optional): 图像预处理的转换函数
        """
        self.image_root = image_root
        self.transform = transform
        self.gt_data = self.importGT(gt_filepath)

    # def importGT(self, gt_filepath):
    #     try:
    #         workbook = openpyxl.load_workbook(gt_filepath)
    #         # 这里可以根据实际需求获取工作表等操作
    #         sheet = workbook.active
    #         data = []
    #         is_header = True  # 标记是否为表头行

    #         for row in sheet.iter_rows(values_only=True):
    #             if is_header:
    #                 is_header = False  # 跳过表头行
    #                 continue
    #             processed_row = []
    #             for cell in row:
    #                 if isinstance(cell, str):
    #                     try:
    #                         cell = float(cell)
    #                     except ValueError:
    #                         pass
    #                 processed_row.append(cell)
    #             data.append(processed_row)
    #         return data
    #             # data.append(row)
    #         # return data
    #     except Exception as e:
    #         print(f"Error reading Excel file: {e}")
    #         return None

    def importGT(self, gt_filepath):
        workbook = openpyxl.load_workbook(gt_filepath)
        sheet = workbook.active
        data = []
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
            if row_idx == 0:  # 跳过表头
                continue
                
            processed_row = []
            for col_idx, cell in enumerate(row):
                # 处理标签列（从第8列开始）
                if col_idx >= 7:  
                    try:
                        val = float(cell) if cell is not None else 0.0  # 空值替换为0
                    except (ValueError, TypeError):
                        print(f"警告：无效标签值 @行{row_idx+1} 列{col_idx+1}，值'{cell}'已替换为0")
                        val = 0.0
                    processed_row.append(val)
                else:
                    processed_row.append(cell)
            data.append(processed_row)
        return data

    # def __len__(self):
    #     """
    #     返回数据集的长度
    #     """
    #     return len(self.gt_data) * 2  # 因为每个病例有左右眼两张图像

    def __len__(self):
        # 实际检查存在的图像数量
        valid_count = 0
        for case in self.gt_data:
            case_id = case[0]
            left_path = os.path.join(self.image_root, f"{case_id}_left.jpg")
            right_path = os.path.join(self.image_root, f"{case_id}_right.jpg")
            if os.path.exists(left_path):
                valid_count +=1
            if os.path.exists(right_path):
                valid_count +=1
        return valid_count

    def __getitem__(self, idx):
        """
        根据索引获取数据集中的样本
        参数:
            idx (int): 样本索引
        返回:
            tuple: 包含图像和标签的元组
        """
        case_index = idx // 2
        is_left = idx % 2 == 0

        # 直接从 gt_data 中获取 case_id
        # case_id = self.gt_data[case_index, 0]
        case_id = int(self.gt_data[case_index][0])  # 修改为使用两个整数索引
        # labels = torch.tensor(self.gt_data[case_index, 1:], dtype=torch.float32)

        # 获取标签数据，标签从第8列开始
        labels = torch.tensor(self.gt_data[case_index][7:], dtype=torch.float32)

        # 根据左右眼情况生成图像文件名
        eye_suffix = "left" if is_left else "right"
        image_filename = f"{case_id}_{eye_suffix}.jpg"
        image_path = os.path.join(self.image_root, image_filename)

        # 读取图像
        image = cv2.imread(image_path)
        if image is None:
            raise FileNotFoundError(f"Image not found: {image_path}")
        image = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)

        if self.transform:
            image = self.transform(image)

        # 尝试将标签数据转换为数值类型
        try:
            label_data = [float(x) for x in self.gt_data[case_index][7:]]
            labels = torch.tensor(label_data, dtype=torch.float32)
        except ValueError as e:
            print(f"Error converting label data at index {case_index}: {e}")
            print(f"Data: {self.gt_data[case_index][7:]}")
            raise

        return image, labels
# class ODIRDataset(torch.utils.data.Dataset):
#     def __init__(self, image_root, gt_filepath, transform=None):
#         self.image_root = image_root
#         self.transform = transform
#         self.gt_data = self.importGT(gt_filepath)

#     def __getitem__(self, case_index):
#         case_id = self.gt_data[case_index][0]
#         image_path = os.path.join(self.image_root, f"{case_id}.jpg")
#         image = Image.open(image_path).convert('RGB')

#         if self.transform:
#             image = self.transform(image)

#         try:
#             # 尝试将标签数据转换为数值类型
#             label_data = [float(x) for x in self.gt_data[case_index][7:]]
#             labels = torch.tensor(label_data, dtype=torch.float32)
#         except ValueError as e:
#             print(f"Error converting label data at index {case_index}: {e}")
#             print(f"Data: {self.gt_data[case_index][7:]}")
#             raise

#         return image, labels

#     def __len__(self):
#         return len(self.gt_data)

#     def importGT(self, gt_filepath):
#         workbook = openpyxl.load_workbook(gt_filepath)
#         sheet = workbook.active
#         data = []
#         is_header = True  # 标记是否为表头行
#         for row in sheet.iter_rows(values_only=True):
#             if is_header:
#                 is_header = False  # 跳过表头行
#                 continue
#             processed_row = []
#             for cell in row:
#                 if isinstance(cell, str):
#                     try:
#                         cell = float(cell)
#                     except ValueError:
#                         pass
#                 processed_row.append(cell)
#             data.append(processed_row)
#         return data

def load_OIA_ODIR_dataset(train_image_root, test_image_root, train_gt_filepath, test_gt_filepath, n_sites, image_size=512, batch_size=64, val_ratio=0.3):
    """
    加载自定义数据集
    参数:
        train_image_root (str): 训练集图像文件所在的根目录
        test_image_root (str): 测试集图像文件所在的根目录
        train_gt_filepath (str): 训练集的真实标签 Excel 文件路径
        test_gt_filepath (str): 测试集的真实标签 Excel 文件路径
        image_size (int): 图像统一缩放尺寸（默认512x512）
        batch_size (int): 数据加载批次大小
        val_ratio (float): 验证集划分比例（默认0.3）
    返回:
        (train_loaders, val_loaders, test_loaders): 训练/验证/测试数据加载器列表
    """
    # 数据预处理配置
    transform = transforms.Compose([
        transforms.ToPILImage(),
        transforms.Resize((image_size, image_size)),  # 统一图像尺寸
        transforms.RandomHorizontalFlip(p=0.5),
        transforms.RandomRotation(10),
        transforms.ToTensor(),
        transforms.Normalize(mean=[0.485, 0.456, 0.406],  # ImageNet标准归一化
                             std=[0.229, 0.224, 0.225])
    ])

    # 加载完整训练集
    full_train = ODIRDataset(train_image_root, train_gt_filepath, transform=transform)
    # 划分训练集到 n 个站点
    total_train_samples = len(full_train)
    full_train_length = len(full_train)
    samples_per_site = total_train_samples // n_sites
    remaining_samples = full_train_length % n_sites

    # 计算每个子集的长度
    lengths = [samples_per_site] * n_sites
    if remaining_samples > 0:
        lengths[-1] += remaining_samples

    print(f"Full train dataset length: {full_train_length}")
    print(f"Samples per site: {samples_per_site}")
    print(f"Remaining samples: {remaining_samples}")
    print(f"Lengths of subsets: {lengths}")
    print(f"Sum of lengths: {sum(lengths)}")

    # 检查长度是否匹配
    if sum(lengths) != full_train_length:
        print("Warning: Sum of lengths does not equal the length of the input dataset!")

    train_datasets = random_split(full_train, lengths)


    # 加载完整测试集
    full_test = ODIRDataset(test_image_root, test_gt_filepath, transform=transform)
    # 划分测试集到 n 个站点
    total_test_samples = len(full_test)
    test_samples_per_site = total_test_samples // n_sites
    test_remaining_samples = total_test_samples % n_sites

    # 计算测试集每个子集的长度
    test_lengths = [test_samples_per_site] * n_sites
    if test_remaining_samples > 0:
        test_lengths[-1] += test_remaining_samples

    print(f"Full test dataset length: {total_test_samples}")
    print(f"Test samples per site: {test_samples_per_site}")
    print(f"Test remaining samples: {test_remaining_samples}")
    print(f"Test lengths of subsets: {test_lengths}")
    print(f"Test sum of lengths: {sum(test_lengths)}")

    # 检查测试集长度是否匹配
    if sum(test_lengths) != total_test_samples:
        print("Warning: Test sum of lengths does not equal the length of the test dataset!")

    test_datasets = random_split(full_test, test_lengths)

    train_loaders = []
    val_loaders = []
    test_loaders = []

    for i in range(n_sites):
        # 为每个站点的训练集划分验证集
        val_size = int(len(train_datasets[i]) * val_ratio)
        train_size = len(train_datasets[i]) - val_size
        train_subset, val_subset = random_split(
            train_datasets[i],
            [train_size, val_size],
            generator=torch.Generator().manual_seed(42)  # 固定随机种子保证可重复性
        )

        # 创建数据加载器
        train_loader = DataLoader(train_subset, batch_size=batch_size,
                                  shuffle=True, num_workers=4, pin_memory=True)
        val_loader = DataLoader(val_subset, batch_size=batch_size,
                                shuffle=False, num_workers=4, pin_memory=True)
        test_loader = DataLoader(test_datasets[i], batch_size=batch_size,
                                 shuffle=False, num_workers=4, pin_memory=True)

        train_loaders.append(train_loader)
        val_loaders.append(val_loader)
        test_loaders.append(test_loader)

    return train_loaders, val_loaders, test_loaders


# 加载eyepac-light-v2数据集使用
def load_custom_dataset(train_root, test_root, n_sites, image_size=224, batch_size=64, val_ratio=0.3):
    """
    加载自定义数据集（按类别文件夹组织）
    参数:
        train_root (str): 训练集根目录路径（包含按类别组织的子文件夹）
        test_root (str): 测试集根目录路径（结构同训练集）
        image_size (int): 图像统一缩放尺寸-默认224x224
        batch_size (int): 数据加载批次大小
        val_ratio (float): 验证集划分比例（默认0.3）
    返回:
        (train_loader, val_loader, test_loader, class_names): 
            训练/验证/测试数据加载器 + 类别名称列表
    """
    # 数据预处理配置
    transform = transforms.Compose([
        transforms.Resize((image_size, image_size)),  # 统一图像尺寸
        transforms.RandomHorizontalFlip(p=0.5),
        transforms.RandomRotation(10),
        transforms.ToTensor(),
        transforms.Normalize(mean=[0.485, 0.456, 0.406],  # ImageNet标准归一化
                          std=[0.229, 0.224, 0.225])
    ])
    
    # 加载完整训练集================================================
    full_train = datasets.ImageFolder(root=train_root, transform=transform)
    # 划分训练集到 n 个站点
    total_train_samples = len(full_train)
    samples_per_site = total_train_samples // n_sites
    # train_datasets = random_split(full_train, [samples_per_site] * n_sites)
    remaining_samples = total_train_samples % n_sites
    # 计算每个子集的长度
    lengths = [samples_per_site] * n_sites
    if remaining_samples > 0:
        lengths[-1] += remaining_samples
    train_datasets = random_split(full_train, lengths)
    print(f"Full train dataset length: {total_train_samples}")
    print(f"Samples per site: {samples_per_site}")
    print(f"Remaining samples: {remaining_samples}")
    print(f"Lengths of subsets: {lengths}")
    print(f"Sum of lengths: {sum(lengths)}")

    # 加载完整测试集==================================================
    full_test = datasets.ImageFolder(root=test_root, transform=transform)
    # 划分测试集到 n 个站点
    total_test_samples = len(full_test)
    test_samples_per_site = total_test_samples // n_sites
    test_remaining_samples = total_test_samples % n_sites  # 防止无法整除
    # 计算测试集每个子集的长度
    test_lengths = [test_samples_per_site] * n_sites
    if test_remaining_samples > 0:
        test_lengths[-1] += test_remaining_samples
    # test_datasets = random_split(full_test, [test_samples_per_site] * n_sites)
    test_datasets = random_split(full_test, test_lengths)
    print(f"Full test dataset length: {total_test_samples}")
    print(f"Test samples per site: {test_samples_per_site}")
    print(f"Test remaining samples: {test_remaining_samples}")
    print(f"Test lengths of subsets: {test_lengths}")
    print(f"Test sum of lengths: {sum(test_lengths)}")

    train_loaders = []
    val_loaders = []
    test_loaders = []
    
    for i in range(n_sites):
        # 为每个站点的训练集划分验证集
        val_size = int(len(train_datasets[i]) * val_ratio)
        train_size = len(train_datasets[i]) - val_size
        train_subset, val_subset = random_split(
            train_datasets[i],
            [train_size, val_size],
            generator=torch.Generator().manual_seed(42)  # 固定随机种子保证可重复性
        )

        # 创建数据加载器
        train_loader = DataLoader(train_subset, batch_size=batch_size,
                                  shuffle=True, num_workers=4, pin_memory=True, drop_last=True)
        val_loader = DataLoader(val_subset, batch_size=batch_size,
                                shuffle=False, num_workers=4, pin_memory=True, drop_last=True)
        test_loader = DataLoader(test_datasets[i], batch_size=batch_size,
                                 shuffle=False, num_workers=4, pin_memory=True,drop_last=True)

        train_loaders.append(train_loader)
        val_loaders.append(val_loader)
        test_loaders.append(test_loader)
    
    # 获取类别名称映射
    class_names = full_train.classes
    
    return train_loaders, val_loaders, test_loaders, class_names


def load_mnist_data(n_sites, batch_size=64):
    """
    加载 MNIST 数据集并分发到 n 个本地站点。
    参数:
        n_sites (int): 本地站点的数量。
        batch_size (int): 每个站点的数据批大小。
    返回:
        train_loaders (list): 每个站点的训练数据加载器列表。
        test_loaders (list): 每个站点的测试数据加载器列表。
    """
    # 定义数据预处理
    transform = transforms.Compose([
        transforms.ToTensor(),
        transforms.Normalize((0.5,), (0.5,))
    ])

    # 加载完整数据集
    full_train_dataset = datasets.MNIST(root='./data', train=True, download=True, transform=transform)
    full_test_dataset = datasets.MNIST(root='./data', train=False, download=True, transform=transform)

    # 计算每个站点的训练数据量
    total_train_samples = len(full_train_dataset)
    samples_per_site = total_train_samples // n_sites

    # 分割训练数据集
    train_datasets = random_split(full_train_dataset, [samples_per_site] * n_sites)

    # 为每个站点创建训练和测试数据加载器
    train_loaders = []
    test_loaders = []

    for i in range(n_sites):
        # 从每个站点的训练集中分离出测试集
        train_subset, test_subset = random_split(train_datasets[i], [int(0.8 * samples_per_site), int(0.2 * samples_per_site)])
        
        # 创建数据加载器
        train_loader = DataLoader(train_subset, batch_size=batch_size, shuffle=True)
        test_loader = DataLoader(test_subset, batch_size=batch_size, shuffle=False)
        
        train_loaders.append(train_loader)
        test_loaders.append(test_loader)

    return train_loaders, test_loaders